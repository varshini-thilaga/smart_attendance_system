import sys
import os
import sqlite3
import pickle
import shutil
import importlib
from datetime import datetime
import face_recognition
import cv2
import numpy as np

from PyQt5.QtWidgets import (
    QApplication, QWidget, QPushButton, QVBoxLayout, QHBoxLayout,
    QLabel, QMainWindow, QMessageBox, QInputDialog, QFileDialog,
    QGroupBox, QGridLayout, QFrame, QProgressDialog
)
from PyQt5.QtGui import QPixmap, QFont, QPainter, QLinearGradient, QColor, QBrush, QIcon
from PyQt5.QtCore import Qt, QSize, QThread, pyqtSignal

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "Face-DataBase")
MODEL_FILE = os.path.join(BASE_DIR, "trained_model.pkl")
DATASET_DIR = os.path.join(BASE_DIR, "dataset")
CROPPED_DIR = os.path.join(BASE_DIR, "Cropped_faces")
PICS_DIR = os.path.join(BASE_DIR, "pics")


def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""CREATE TABLE IF NOT EXISTS Students (
        ID TEXT PRIMARY KEY, Name TEXT, Roll TEXT, personID TEXT)""")
    conn.execute("""CREATE TABLE IF NOT EXISTS TrainLog (trained_at TEXT)""")
    conn.commit()
    conn.close()


class TrainThread(QThread):
    finished = pyqtSignal(int)
    error = pyqtSignal(str)

    def run(self):
        try:
            known_encodings, known_ids = [], []
            for folder in os.listdir(DATASET_DIR):
                folder_path = os.path.join(DATASET_DIR, folder)
                if not os.path.isdir(folder_path):
                    continue
                student_id = folder.replace("user", "")
                for filename in os.listdir(folder_path):
                    if not filename.lower().endswith((".jpg", ".jpeg", ".png")):
                        continue
                    img = face_recognition.load_image_file(os.path.join(folder_path, filename))
                    encs = face_recognition.face_encodings(img)
                    if encs:
                        known_encodings.append(encs[0])
                        known_ids.append(student_id)

            with open(MODEL_FILE, "wb") as f:
                pickle.dump({"encodings": known_encodings, "ids": known_ids}, f)

            conn = sqlite3.connect(DB_PATH)
            conn.execute("DELETE FROM TrainLog")
            conn.execute("INSERT INTO TrainLog VALUES (?)", (datetime.now().isoformat(),))
            conn.commit()
            conn.close()
            self.finished.emit(len(known_encodings))
        except Exception as e:
            self.error.emit(str(e))


class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        init_db()
        self.setWindowTitle("AttendEase - Smart Attendance System")
        self.setMinimumSize(900, 580)
        self.setWindowIcon(QIcon(os.path.join(BASE_DIR, "home.png")))
        self._build_ui()

    def paintEvent(self, event):
        painter = QPainter(self)
        gradient = QLinearGradient(0, 0, self.width(), self.height())
        gradient.setColorAt(0.0, QColor("#0f2027"))
        gradient.setColorAt(0.5, QColor("#203a43"))
        gradient.setColorAt(1.0, QColor("#2c5364"))
        painter.fillRect(self.rect(), QBrush(gradient))

    def _build_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(40, 30, 40, 30)
        main_layout.setSpacing(20)

        # Title
        title = QLabel("AttendEase")
        title.setAlignment(Qt.AlignCenter)
        title.setFont(QFont("Segoe UI", 32, QFont.Bold))
        title.setStyleSheet("color: #00d4ff; letter-spacing: 4px;")
        main_layout.addWidget(title)

        subtitle = QLabel("Smart Attendance Management System")
        subtitle.setAlignment(Qt.AlignCenter)
        subtitle.setFont(QFont("Segoe UI", 11))
        subtitle.setStyleSheet("color: #aaaaaa;")
        main_layout.addWidget(subtitle)

        # Divider
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setStyleSheet("color: #2c5364;")
        main_layout.addWidget(line)

        # Buttons grid
        grid = QGridLayout()
        grid.setSpacing(16)

        buttons = [
            ("Enroll Student",     "👤", "#1a73e8", self.enroll),
            ("Train Model",        "🧠", "#0f9d58", self.train),
            ("Mark Attendance",    "📸", "#e37400", self.mark_attendance),
            ("View Attendance",    "📊", "#9c27b0", self.view_attendance),
            ("Edit Timetable",     "📅", "#00838f", self.timetable),
            ("Model Status",       "ℹ️",  "#546e7a", self.status),
        ]

        for idx, (label, icon, color, handler) in enumerate(buttons):
            btn = QPushButton(f"  {icon}  {label}")
            btn.setMinimumHeight(70)
            btn.setFont(QFont("Segoe UI", 12, QFont.Bold))
            btn.setCursor(Qt.PointingHandCursor)
            btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {color};
                    color: white;
                    border-radius: 10px;
                    border: none;
                    text-align: left;
                    padding-left: 20px;
                }}
                QPushButton:hover {{
                    background-color: white;
                    color: {color};
                    border: 2px solid {color};
                }}
                QPushButton:pressed {{
                    opacity: 0.8;
                }}
            """)
            btn.clicked.connect(handler)
            grid.addWidget(btn, idx // 3, idx % 3)

        main_layout.addLayout(grid)

        # Status bar
        self.status_label = QLabel("Ready")
        self.status_label.setAlignment(Qt.AlignCenter)
        self.status_label.setFont(QFont("Segoe UI", 9))
        self.status_label.setStyleSheet("color: #aaaaaa; padding: 6px;")
        main_layout.addWidget(self.status_label)

    def set_status(self, msg):
        self.status_label.setText(msg)

    # ── Enroll Student ──────────────────────────────────────────────
    def enroll(self):
        name, ok1 = QInputDialog.getText(self, "Enroll Student", "Enter student name:")
        if not ok1 or not name.strip():
            return
        roll, ok2 = QInputDialog.getText(self, "Enroll Student", "Enter roll number:")
        if not ok2 or not roll.strip():
            return

        student_id = roll.strip()[-3:]

        files, _ = QFileDialog.getOpenFileNames(
            self, "Select Face Photos for " + name, "",
            "Images (*.jpg *.jpeg *.png)")
        if not files:
            QMessageBox.warning(self, "No Images", "Please select at least one face photo.")
            return

        # Save to DB
        conn = sqlite3.connect(DB_PATH)
        cur = conn.execute("SELECT * FROM Students WHERE ID=?", (student_id,))
        if cur.fetchone():
            conn.execute("UPDATE Students SET Name=?, Roll=? WHERE ID=?", (name, roll, student_id))
        else:
            conn.execute("INSERT INTO Students(ID,Name,Roll) VALUES(?,?,?)", (student_id, name, roll))
        conn.commit()
        conn.close()

        # Save images (handle non-English paths)
        folder = os.path.join(DATASET_DIR, "user" + student_id)
        os.makedirs(folder, exist_ok=True)
        saved = 0
        for i, f in enumerate(files[:20]):
            img_array = np.fromfile(f, dtype=np.uint8)
            img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
            if img is not None:
                cv2.imwrite(os.path.join(folder, f"User.{student_id}.{i+1}.jpg"), img)
                saved += 1

        QMessageBox.information(self, "Enrolled",
            f"✅ {name} enrolled successfully!\n{saved} photo(s) saved.\n\nClick 'Train Model' to update the model.")
        self.set_status(f"Enrolled: {name} ({roll})")

    # ── Train Model ─────────────────────────────────────────────────
    def train(self):
        if not os.path.exists(DATASET_DIR) or not os.listdir(DATASET_DIR):
            QMessageBox.warning(self, "No Data", "No students enrolled yet. Please enroll students first.")
            return

        if os.path.exists(MODEL_FILE):
            reply = QMessageBox.question(self, "Retrain?",
                "A trained model already exists.\nOnly retrain if you enrolled new students.\n\nRetrain now?",
                QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.No:
                return

        self.set_status("Training model... please wait.")
        self.progress = QProgressDialog("Training face recognition model...", None, 0, 0, self)
        self.progress.setWindowTitle("Training")
        self.progress.setWindowModality(Qt.WindowModal)
        self.progress.show()

        self.train_thread = TrainThread()
        self.train_thread.finished.connect(self._train_done)
        self.train_thread.error.connect(self._train_error)
        self.train_thread.start()

    def _train_done(self, count):
        self.progress.close()
        QMessageBox.information(self, "Training Complete", f"✅ Model trained successfully!\n{count} face(s) encoded.")
        self.set_status(f"Model trained — {count} faces encoded.")

    def _train_error(self, msg):
        self.progress.close()
        QMessageBox.critical(self, "Training Failed", f"Error: {msg}")
        self.set_status("Training failed.")

    # ── Mark Attendance ─────────────────────────────────────────────
    def mark_attendance(self):
        if not os.path.exists(MODEL_FILE):
            QMessageBox.warning(self, "No Model", "Please train the model first.")
            return

        file, _ = QFileDialog.getOpenFileName(
            self, "Select Class Photo", "", "Images (*.jpg *.jpeg *.png)")
        if not file:
            return

        subject, ok = QInputDialog.getText(self, "Subject", "Enter subject name:")
        if not ok or not subject.strip():
            return

        os.makedirs(CROPPED_DIR, exist_ok=True)
        os.makedirs(PICS_DIR, exist_ok=True)

        # Clear old crops
        for f in os.listdir(CROPPED_DIR):
            os.remove(os.path.join(CROPPED_DIR, f))

        # Detect and crop faces (handle non-English paths)
        import dlib
        detector = dlib.get_frontal_face_detector()
        img_array = np.fromfile(file, dtype=np.uint8)
        img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
        if img is None:
            QMessageBox.critical(self, "Error", "Could not read the selected image.")
            return
        cv2.imwrite(os.path.join(PICS_DIR, "framee1.jpg"), img)
        dets = detector(img, 1)

        if not dets:
            QMessageBox.warning(self, "No Faces", "No faces detected in the selected image.")
            return

        for i, d in enumerate(dets):
            cv2.imwrite(os.path.join(CROPPED_DIR, f"face{i+1}.jpg"),
                        img[d.top():d.bottom(), d.left():d.right()])

        # Identify faces
        with open(MODEL_FILE, "rb") as f:
            data = pickle.load(f)
        known_encodings = data["encodings"]
        known_ids = data["ids"]

        conn = sqlite3.connect(DB_PATH)
        present_ids = set()

        for fname in os.listdir(CROPPED_DIR):
            if not fname.lower().endswith(".jpg"):
                continue
            face_img = face_recognition.load_image_file(os.path.join(CROPPED_DIR, fname))
            encs = face_recognition.face_encodings(face_img)
            if not encs:
                continue
            if not known_encodings:
                continue
            distances = face_recognition.face_distance(known_encodings, encs[0])
            best = int(np.argmin(distances))
            print(f"{fname} -> best match: ID={known_ids[best]}, distance={distances[best]:.3f}")
            if distances[best] < 0.65:  # relaxed tolerance
                present_ids.add(known_ids[best])

        # Get all students from DB — normalize all IDs to string for comparison
        cursor = conn.execute("SELECT Roll, Name, ID FROM Students ORDER BY Roll ASC")
        students = cursor.fetchall()
        present_ids_str = set(str(pid) for pid in present_ids)
        print(f"Present IDs from model: {present_ids_str}")
        print(f"Students in DB: {[(str(s[2]), s[1]) for s in students]}")

        # Write to Excel
        currentdate = datetime.now().strftime("%d_%m_%y")
        out_dir = os.path.join(BASE_DIR, currentdate, subject.strip())
        os.makedirs(out_dir, exist_ok=True)
        out_file = os.path.join(out_dir, f"{subject.strip()}_{currentdate}.xlsx")

        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = subject.strip()[:31]
        ws.append(["Roll Number", "Name", "Status"])
        present_names = []
        for roll, name, sid in students:
            status_val = "Present" if str(sid) in present_ids_str else "Absent"
            ws.append([roll, name, status_val])
            if str(sid) in present_ids_str:
                present_names.append(name)

        wb.save(out_file)
        conn.close()

        QMessageBox.information(self, "Attendance Marked",
            f"✅ Attendance marked for {subject}!\n\n"
            f"Detected: {len(dets)} faces\n"
            f"Recognized: {len(present_ids)} students\n\n"
            f"Present: {', '.join(present_names) if present_names else 'None'}\n\n"
            f"Saved to: {out_file}")
        self.set_status(f"Attendance marked — {subject} — {len(present_ids)} present")

    # ── View Attendance ─────────────────────────────────────────────
    def view_attendance(self):
        self.attendance_window = AttendanceViewer()
        self.attendance_window.show()

    # ── Timetable ───────────────────────────────────────────────────
    def timetable(self):
        os.system(f'python "{os.path.join(BASE_DIR, "timetable", "time.py")}"')

    # ── Model Status ────────────────────────────────────────────────
    def status(self):
        import get_status
        importlib.reload(get_status)
        res = get_status.status()
        if "not trained" in res.lower():
            QMessageBox.information(self, "Model Status", "⚠️ Model has not been trained yet.")
        else:
            try:
                dt = datetime.fromisoformat(res)
                formatted = dt.strftime("%d-%m-%Y at %H:%M:%S")
                QMessageBox.information(self, "Model Status", f"✅ Model last trained on:\n{formatted}")
            except:
                QMessageBox.information(self, "Model Status", res)
        self.set_status("Status checked.")


class AttendanceViewer(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Attendance Records")
        self.setMinimumSize(800, 500)
        self._build_ui()
        self._load_records()

    def paintEvent(self, event):
        painter = QPainter(self)
        gradient = QLinearGradient(0, 0, self.width(), self.height())
        gradient.setColorAt(0.0, QColor("#0f2027"))
        gradient.setColorAt(1.0, QColor("#2c5364"))
        painter.fillRect(self.rect(), QBrush(gradient))

    def _build_ui(self):
        from PyQt5.QtWidgets import QTableWidget, QTableWidgetItem, QHeaderView, QComboBox
        self.QTableWidget = QTableWidget
        self.QTableWidgetItem = QTableWidgetItem
        self.QHeaderView = QHeaderView

        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        title = QLabel("📊 Attendance Records")
        title.setFont(QFont("Segoe UI", 16, QFont.Bold))
        title.setStyleSheet("color: #00d4ff;")
        layout.addWidget(title)

        # Filter row
        filter_layout = QHBoxLayout()
        self.date_combo = QComboBox()
        self.date_combo.setStyleSheet("background:#203a43; color:white; padding:6px; border-radius:6px;")
        self.date_combo.currentTextChanged.connect(self._load_records)
        filter_layout.addWidget(QLabel("Date:"))
        filter_layout.addWidget(self.date_combo)

        self.subject_combo = QComboBox()
        self.subject_combo.setStyleSheet("background:#203a43; color:white; padding:6px; border-radius:6px;")
        self.subject_combo.currentTextChanged.connect(self._load_records)
        filter_layout.addWidget(QLabel("Subject:"))
        filter_layout.addWidget(self.subject_combo)
        filter_layout.addStretch()

        export_btn = QPushButton("📥 Export to Excel")
        export_btn.setStyleSheet("background:#1a73e8; color:white; padding:8px 16px; border-radius:6px; border:none;")
        export_btn.clicked.connect(self._export)
        filter_layout.addWidget(export_btn)

        for lbl in self.findChildren(QLabel):
            lbl.setStyleSheet("color: #aaaaaa;")
        layout.addLayout(filter_layout)

        self.table = QTableWidget()
        self.table.setStyleSheet("""
            QTableWidget { background:#1a2a35; color:white; gridline-color:#2c5364; border:none; }
            QHeaderView::section { background:#203a43; color:#00d4ff; padding:8px; border:none; font-weight:bold; }
            QTableWidget::item:selected { background:#1a73e8; }
        """)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        layout.addWidget(self.table)

        self.summary_label = QLabel("")
        self.summary_label.setStyleSheet("color:#aaaaaa; font-size:10pt;")
        layout.addWidget(self.summary_label)

        self._populate_combos()

    def _find_attendance_dirs(self):
        dirs = []
        for d in sorted(os.listdir(BASE_DIR)):
            d_path = os.path.join(BASE_DIR, d)
            # match format dd_mm_yy
            if os.path.isdir(d_path) and len(d) == 8 and d.count('_') == 2:
                try:
                    datetime.strptime(d, "%d_%m_%y")
                    dirs.append((d, d_path))
                except:
                    pass
        return dirs

    def _populate_combos(self):
        self.date_combo.blockSignals(True)
        self.subject_combo.blockSignals(True)
        self.date_combo.clear()
        self.subject_combo.clear()
        self.date_combo.addItem("All Dates")
        self.subject_combo.addItem("All Subjects")

        for d, d_path in self._find_attendance_dirs():
            self.date_combo.addItem(d)
            for s in os.listdir(d_path):
                if os.path.isdir(os.path.join(d_path, s)) and self.subject_combo.findText(s) == -1:
                    self.subject_combo.addItem(s)

        self.date_combo.blockSignals(False)
        self.subject_combo.blockSignals(False)

    def _load_records(self):
        sel_date = self.date_combo.currentText()
        sel_subject = self.subject_combo.currentText()

        from openpyxl import load_workbook
        all_rows = []

        for d, d_path in self._find_attendance_dirs():
            if sel_date != "All Dates" and d != sel_date:
                continue
            for subject in os.listdir(d_path):
                s_path = os.path.join(d_path, subject)
                if not os.path.isdir(s_path):
                    continue
                if sel_subject != "All Subjects" and subject != sel_subject:
                    continue
                for f in os.listdir(s_path):
                    if f.endswith(".xlsx"):
                        try:
                            wb = load_workbook(os.path.join(s_path, f))
                            ws = wb.active
                            for row in ws.iter_rows(min_row=2, values_only=True):
                                if row and row[0]:
                                    all_rows.append((d, subject, row[0], row[1], row[2] or "Absent"))
                        except Exception as e:
                            print(f"Error reading {f}: {e}")

        self.table.setRowCount(len(all_rows))
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["Date", "Subject", "Roll No", "Name", "Status"])

        present = 0
        for i, (date, subject, roll, name, status) in enumerate(all_rows):
            for j, val in enumerate([date, subject, roll, name, status]):
                item = self.QTableWidgetItem(str(val) if val else "")
                item.setTextAlignment(Qt.AlignCenter)
                if str(status) == "Present":
                    item.setForeground(QColor("#0f9d58"))
                elif str(status) == "Absent":
                    item.setForeground(QColor("#e53935"))
                self.table.setItem(i, j, item)
            if status == "Present":
                present += 1

        total = len(all_rows)
        self.summary_label.setText(
            f"Total Records: {total}  |  Present: {present}  |  Absent: {total - present}")

    def _export(self):
        from openpyxl import Workbook
        path, _ = QFileDialog.getSaveFileName(self, "Save Excel", "", "Excel (*.xlsx)")
        if not path:
            return
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Subject", "Roll No", "Name", "Status"])
        for row in range(self.table.rowCount()):
            ws.append([self.table.item(row, col).text() for col in range(5)])
        wb.save(path)
        QMessageBox.information(self, "Exported", f"Saved to {path}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
