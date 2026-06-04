from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
import sqlite3
import pandas as pd
from datetime import datetime, date

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "Face-DataBase")

conn = sqlite3.connect(DB_PATH)
c = conn.cursor()

now = datetime.now()
current_day = date.today().weekday()
currentdate = now.strftime("%d_%m_%y")

timetable_path = os.path.join(BASE_DIR, "time_table.csv")
if not os.path.exists(timetable_path):
    print("time_table.csv not found.")
    exit()

data = pd.read_csv(timetable_path)
periods = data.values[current_day + 1][1:]

out_dir = os.path.join(BASE_DIR, currentdate)
os.makedirs(out_dir, exist_ok=True)

for subject in periods:
    subject_dir = os.path.join(out_dir, str(subject))
    os.makedirs(subject_dir, exist_ok=True)
    out_file = os.path.join(subject_dir, f"{subject}_{currentdate}.xlsx")

    if not os.path.exists(out_file):
        wb = Workbook()
        ws = wb.active
        ws.title = str(subject)
        ws.append(["Roll Number", "Name", "Status"])
        c.execute("SELECT Roll, Name FROM Students ORDER BY Roll ASC")
        for row in c.fetchall():
            ws.append([row[0], row[1], "Absent"])
        wb.save(out_file)
        print(f"Created: {out_file}")

conn.close()
