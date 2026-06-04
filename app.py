from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file
import os
from datetime import datetime
from functools import wraps
from database import get_conn, hash_password, init_db
from face_engine import train_model, save_student_images, detect_and_identify
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = "attendease_secret_2024"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
EXCEL_DIR = os.path.join(BASE_DIR, "excel_reports")
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(EXCEL_DIR, exist_ok=True)


# ── Auth Decorator ───────────────────────────────────────────────────
def login_required(roles=None):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if "user_id" not in session:
                return redirect(url_for("login"))
            if roles and session.get("role") not in roles:
                return redirect(url_for("dashboard"))
            return f(*args, **kwargs)
        return wrapper
    return decorator


# ── Helpers ───────────────────────────────────────────────────────────
def get_daily_summary(class_name, section, date_str, conn):
    """Return dict with fn_present, fn_absent, an_present, an_absent for a class+date."""
    total = conn.execute(
        "SELECT COUNT(*) as c FROM Students WHERE class_name=? AND section=?",
        (class_name, section)
    ).fetchone()["c"]
    row = conn.execute("""
        SELECT
            SUM(CASE WHEN a.fn_status='P' THEN 1 ELSE 0 END) as fn_p,
            SUM(CASE WHEN a.fn_status='Ab' THEN 1 ELSE 0 END) as fn_ab,
            SUM(CASE WHEN a.an_status='P' THEN 1 ELSE 0 END) as an_p,
            SUM(CASE WHEN a.an_status='Ab' THEN 1 ELSE 0 END) as an_ab
        FROM Attendance a JOIN Students s ON a.student_id=s.id
        WHERE s.class_name=? AND s.section=? AND a.date=?
    """, (class_name, section, date_str)).fetchone()
    return {
        "total": total,
        "fn_present": row["fn_p"] or 0,
        "fn_absent": row["fn_ab"] or 0,
        "an_present": row["an_p"] or 0,
        "an_absent": row["an_ab"] or 0,
    }


# ── Excel Helper ─────────────────────────────────────────────────────
def update_excel(class_name, section, conn):
    """Rebuild the full Excel sheet from DB — includes Total Present / Total Absent summary rows."""
    excel_path = os.path.join(EXCEL_DIR, f"{class_name}_{section}.xlsx")
    students = conn.execute(
        "SELECT id, name, roll_no FROM Students WHERE class_name=? AND section=? ORDER BY name ASC",
        (class_name, section)
    ).fetchall()
    # Collect all (date, session_type) pairs that exist
    sessions = conn.execute("""
        SELECT DISTINCT a.date,
            CASE WHEN a.fn_status IS NOT NULL THEN 'FN' END as stype
        FROM Attendance a JOIN Students s ON a.student_id=s.id
        WHERE s.class_name=? AND s.section=? AND a.fn_status IS NOT NULL
        UNION
        SELECT DISTINCT a.date,
            CASE WHEN a.an_status IS NOT NULL THEN 'AN' END
        FROM Attendance a JOIN Students s ON a.student_id=s.id
        WHERE s.class_name=? AND s.section=? AND a.an_status IS NOT NULL
        ORDER BY 1, 2
    """, (class_name, section, class_name, section)).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = f"{class_name}-{section}"

    # Build header row: Roll No | Name | date FN | date AN | ...
    header = ["Roll No", "Name"]
    session_cols = []  # list of (date, stype, col_idx)
    for s in sessions:
        header.append(f"{s[0]} {s[1]}")
        session_cols.append((s[0], s[1], len(header)))
    header += ["Total Present", "Total Absent"]
    total_present_col = len(header) - 1
    total_absent_col = len(header)
    ws.append(header)

    # Style header
    for col in range(1, len(header) + 1):
        c = ws.cell(1, col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1565C0")
        c.alignment = Alignment(horizontal="center")

    # Fetch all attendance keyed by (student_id, date)
    att_rows = conn.execute("""
        SELECT a.student_id, a.date, a.fn_status, a.an_status
        FROM Attendance a JOIN Students s ON a.student_id=s.id
        WHERE s.class_name=? AND s.section=?
    """, (class_name, section)).fetchall()
    att_map = {(r["student_id"], r["date"]): r for r in att_rows}

    for row_idx, student in enumerate(students, start=2):
        ws.cell(row_idx, 1, student["roll_no"]).alignment = Alignment(horizontal="center")
        ws.cell(row_idx, 2, student["name"])
        p_count = 0
        ab_count = 0
        for (date_val, stype, col_idx) in session_cols:
            att = att_map.get((student["id"], date_val))
            if att:
                status = att["fn_status"] if stype == "FN" else att["an_status"]
            else:
                status = "Ab"
            cell = ws.cell(row_idx, col_idx, status)
            cell.fill = PatternFill("solid", fgColor="C8E6C9" if status == "P" else "FFCDD2")
            cell.alignment = Alignment(horizontal="center")
            if status == "P":
                p_count += 1
            else:
                ab_count += 1
        # Summary columns
        pc = ws.cell(row_idx, total_present_col, p_count)
        pc.fill = PatternFill("solid", fgColor="DBEAFE")
        pc.font = Font(bold=True)
        pc.alignment = Alignment(horizontal="center")
        ac = ws.cell(row_idx, total_absent_col, ab_count)
        ac.fill = PatternFill("solid", fgColor="FEF3C7")
        ac.font = Font(bold=True)
        ac.alignment = Alignment(horizontal="center")

    # Summary footer rows
    footer_row = len(students) + 2
    ws.cell(footer_row, 1, "SUMMARY").font = Font(bold=True)
    ws.cell(footer_row, 2, "Total Present").font = Font(bold=True, color="15803D")
    ws.cell(footer_row + 1, 2, "Total Absent").font = Font(bold=True, color="B91C1C")
    for (date_val, stype, col_idx) in session_cols:
        p = sum(
            1 for st in students
            if att_map.get((st["id"], date_val)) and
               (att_map[(st["id"], date_val)]["fn_status" if stype == "FN" else "an_status"]) == "P"
        )
        ab = len(students) - p
        pc = ws.cell(footer_row, col_idx, p)
        pc.fill = PatternFill("solid", fgColor="DCFCE7")
        pc.font = Font(bold=True)
        pc.alignment = Alignment(horizontal="center")
        ac = ws.cell(footer_row + 1, col_idx, ab)
        ac.fill = PatternFill("solid", fgColor="FEE2E2")
        ac.font = Font(bold=True)
        ac.alignment = Alignment(horizontal="center")

    wb.save(excel_path)
    return excel_path


# ── Auth Routes ──────────────────────────────────────────────────────
@app.route("/")
def index():
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if "user_id" in session:
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "").strip()
        conn = get_conn()
        user = conn.execute(
            "SELECT * FROM Users WHERE email=? AND password=?",
            (email, hash_password(password))
        ).fetchone()
        conn.close()
        if user:
            session.update({
                "user_id": user["id"], "role": user["role"],
                "full_name": user["full_name"],
                "class_name": user["class_name"] or "",
                "section": user["section"] or "",
                "year": user["year"] or str(datetime.now().year)
            })
            return redirect(url_for("dashboard"))
        return render_template("login.html", error="Invalid email or password.")
    return render_template("login.html")


@app.route("/signup", methods=["GET", "POST"])
def signup():
    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "").strip()
        role = request.form.get("role", "class_teacher")
        class_name = request.form.get("class_name", "").strip()
        section = request.form.get("section", "").strip()
        try:
            conn = get_conn()
            conn.execute(
                "INSERT INTO Users(full_name,email,password,role,class_name,section) VALUES(?,?,?,?,?,?)",
                (full_name, email, hash_password(password), role, class_name, section)
            )
            conn.commit()
            conn.close()
            return redirect(url_for("login"))
        except:
            return render_template("signup.html", error="Email already exists.")
    return render_template("signup.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/dashboard")
def dashboard():
    if "user_id" not in session:
        return redirect(url_for("login"))
    if session.get("role") == "admin":
        return redirect(url_for("admin_dashboard"))
    return redirect(url_for("ct_dashboard"))


# ── Admin Portal ─────────────────────────────────────────────────────
@app.route("/admin")
@login_required(["admin"])
def admin_dashboard():
    conn = get_conn()
    classes = conn.execute(
        "SELECT class_name, section, COUNT(*) as student_count FROM Students GROUP BY class_name, section ORDER BY class_name, section"
    ).fetchall()
    s_count = conn.execute("SELECT COUNT(*) as c FROM Students").fetchone()["c"]
    a_count = conn.execute("SELECT COUNT(*) as c FROM Attendance").fetchone()["c"]
    recent = conn.execute("""
        SELECT a.date, a.fn_status, a.an_status, s.name, s.class_name, s.section
        FROM Attendance a JOIN Students s ON a.student_id=s.id
        ORDER BY a.date DESC, a.id DESC LIMIT 8
    """).fetchall()
    today = datetime.now().strftime("%Y-%m-%d")
    class_summaries = []
    for c in classes:
        s = get_daily_summary(c["class_name"], c["section"], today, conn)
        class_summaries.append({
            "class_name": c["class_name"], "section": c["section"],
            "student_count": c["student_count"], **s
        })
    conn.close()
    return render_template("admin/dashboard.html", classes=classes, class_summaries=class_summaries,
                           s_count=s_count, a_count=a_count, recent=recent, today=today)


@app.route("/admin/classes")
@login_required(["admin"])
def admin_classes():
    conn = get_conn()
    classes = conn.execute(
        "SELECT class_name, section, COUNT(*) as student_count FROM Students GROUP BY class_name, section ORDER BY class_name, section"
    ).fetchall()
    conn.close()
    return render_template("admin/classes.html", classes=classes)


@app.route("/admin/classes/<class_name>/<section>")
@login_required(["admin"])
def admin_class_detail(class_name, section):
    conn = get_conn()
    students = conn.execute(
        "SELECT * FROM Students WHERE class_name=? AND section=? ORDER BY name",
        (class_name, section)
    ).fetchall()
    records = conn.execute("""
        SELECT a.date, a.fn_status, a.an_status, s.name, s.roll_no
        FROM Attendance a JOIN Students s ON a.student_id=s.id
        WHERE s.class_name=? AND s.section=?
        ORDER BY s.name ASC, a.date DESC
    """, (class_name, section)).fetchall()
    # Date-wise summary
    dates = conn.execute("""
        SELECT DISTINCT a.date FROM Attendance a JOIN Students s ON a.student_id=s.id
        WHERE s.class_name=? AND s.section=? ORDER BY a.date DESC
    """, (class_name, section)).fetchall()
    date_summaries = [get_daily_summary(class_name, section, d["date"], conn) | {"date": d["date"]} for d in dates]
    excel_exists = os.path.exists(os.path.join(EXCEL_DIR, f"{class_name}_{section}.xlsx"))
    conn.close()
    return render_template("admin/class_detail.html", class_name=class_name, section=section,
                           students=students, records=records, excel_exists=excel_exists,
                           date_summaries=date_summaries)


@app.route("/admin/attendance")
@login_required(["admin"])
def admin_attendance():
    conn = get_conn()
    class_filter = request.args.get("class", "")
    date_filter = request.args.get("date", "")
    query = """
        SELECT a.*, s.name, s.roll_no, s.class_name, s.section
        FROM Attendance a JOIN Students s ON a.student_id=s.id
        WHERE 1=1
    """
    params = []
    if class_filter:
        query += " AND s.class_name=?"
        params.append(class_filter)
    if date_filter:
        query += " AND a.date=?"
        params.append(date_filter)
    query += " ORDER BY a.date DESC, s.name"
    records = conn.execute(query, params).fetchall()
    all_classes = conn.execute("SELECT DISTINCT class_name FROM Students ORDER BY class_name").fetchall()
    conn.close()
    return render_template("admin/attendance.html", records=records,
                           all_classes=all_classes, class_filter=class_filter, date_filter=date_filter)


@app.route("/admin/students")
@login_required(["admin"])
def admin_students():
    conn = get_conn()
    students = conn.execute("SELECT * FROM Students ORDER BY class_name, section, name").fetchall()
    conn.close()
    return render_template("admin/students.html", students=students)


@app.route("/admin/excel/<class_name>/<section>")
@login_required(["admin"])
def admin_excel(class_name, section):
    path = os.path.join(EXCEL_DIR, f"{class_name}_{section}.xlsx")
    if os.path.exists(path):
        return send_file(path, as_attachment=True)
    return "No Excel file found.", 404


# ── Class Teacher Portal ─────────────────────────────────────────────
@app.route("/ct")
@login_required(["class_teacher"])
def ct_dashboard():
    conn = get_conn()
    class_name = session["class_name"]
    section = session["section"]
    s_count = conn.execute(
        "SELECT COUNT(*) as c FROM Students WHERE class_name=? AND section=?",
        (class_name, section)
    ).fetchone()["c"]
    a_count = conn.execute(
        "SELECT COUNT(*) as c FROM Attendance a JOIN Students s ON a.student_id=s.id WHERE s.class_name=? AND s.section=?",
        (class_name, section)
    ).fetchone()["c"]
    recent = conn.execute("""
        SELECT a.date, a.fn_status, a.an_status, s.name, s.roll_no
        FROM Attendance a JOIN Students s ON a.student_id=s.id
        WHERE s.class_name=? AND s.section=?
        ORDER BY a.date DESC LIMIT 10
    """, (class_name, section)).fetchall()
    today = datetime.now().strftime("%Y-%m-%d")
    today_summary = get_daily_summary(class_name, section, today, conn)
    conn.close()
    excel_exists = os.path.exists(os.path.join(EXCEL_DIR, f"{class_name}_{section}.xlsx"))
    return render_template("ct/dashboard.html", s_count=s_count, a_count=a_count,
                           recent=recent, excel_exists=excel_exists, today_summary=today_summary, today=today)


# Enroll
@app.route("/ct/enroll", methods=["GET", "POST"])
@login_required(["class_teacher"])
def ct_enroll():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        roll_no = request.form.get("roll_no", "").strip()
        class_name = session["class_name"]
        section = session["section"]
        photos = request.files.getlist("photos")
        if not name or not roll_no:
            return render_template("ct/enroll.html", error="Name and roll number required.")
        conn = get_conn()
        try:
            conn.execute(
                "INSERT OR REPLACE INTO Students(name,roll_no,class_name,section,teacher_id) VALUES(?,?,?,?,?)",
                (name, roll_no, class_name, section, session["user_id"])
            )
            conn.commit()
            student = conn.execute(
                "SELECT id FROM Students WHERE roll_no=? AND class_name=? AND section=?",
                (roll_no, class_name, section)
            ).fetchone()
            conn.close()
            saved = save_student_images(student["id"], session.get("year", str(datetime.now().year)), class_name, section, photos)
            return render_template("ct/enroll.html",
                success=f"✅ {name} enrolled with {saved} photo(s). Train model to apply.")
        except Exception as e:
            conn.close()
            return render_template("ct/enroll.html", error=str(e))
    conn = get_conn()
    students = conn.execute(
        "SELECT * FROM Students WHERE class_name=? AND section=? ORDER BY name",
        (session["class_name"], session["section"])
    ).fetchall()
    conn.close()
    return render_template("ct/enroll.html", students=students)


@app.route("/ct/delete_student/<int:sid>", methods=["POST"])
@login_required(["class_teacher"])
def ct_delete_student(sid):
    conn = get_conn()
    conn.execute("DELETE FROM Students WHERE id=?", (sid,))
    conn.execute("DELETE FROM Attendance WHERE student_id=?", (sid,))
    conn.commit()
    conn.close()
    return redirect(url_for("ct_enroll"))


# Train
@app.route("/ct/train", methods=["GET", "POST"])
@login_required(["class_teacher"])
def ct_train():
    if request.method == "POST":
        class_name = session["class_name"]
        section = session["section"]
        year = session.get("year", str(datetime.now().year))
        count = train_model(year, class_name, section)
        conn = get_conn()
        conn.execute("INSERT INTO TrainLog(year,class_name,section) VALUES(?,?,?)", (year, class_name, section))
        conn.commit()
        conn.close()
        return render_template("ct/train.html", success=f"✅ Model trained with {count} face(s).", count=count)
    conn = get_conn()
    last = conn.execute(
        "SELECT trained_at FROM TrainLog WHERE year=? AND class_name=? AND section=? ORDER BY id DESC LIMIT 1",
        (session.get("year", str(datetime.now().year)), session["class_name"], session["section"])
    ).fetchone()
    conn.close()
    return render_template("ct/train.html", last_trained=last["trained_at"] if last else None)


# Mark Attendance
@app.route("/ct/mark", methods=["GET", "POST"])
@login_required(["class_teacher"])
def ct_mark():
    if request.method == "POST":
        session_type = request.form.get("session_type", "FN")
        date_str = request.form.get("date", datetime.now().strftime("%Y-%m-%d"))
        photo = request.files.get("photo")
        class_name = session["class_name"]
        section = session["section"]

        if not photo or not photo.filename:
            return render_template("ct/mark.html", error="Please upload a class photo.")

        img_path = os.path.join(UPLOAD_DIR, secure_filename(photo.filename))
        photo.save(img_path)

        year = session.get("year", str(datetime.now().year))
        present_ids, detected, err = detect_and_identify(img_path, year, class_name, section)
        if err:
            return render_template("ct/mark.html", error=err)

        conn = get_conn()
        students = conn.execute(
            "SELECT * FROM Students WHERE class_name=? AND section=? ORDER BY name",
            (class_name, section)
        ).fetchall()

        present_ids_str = [str(p) for p in present_ids]
        present_names = []

        for s in students:
            status = "P" if str(s["id"]) in present_ids_str else "Ab"
            if session_type == "FN":
                conn.execute("""
                    INSERT INTO Attendance(student_id,date,fn_status,marked_by)
                    VALUES(?,?,?,?)
                    ON CONFLICT(student_id,date) DO UPDATE SET fn_status=excluded.fn_status
                """, (s["id"], date_str, status, session["user_id"]))
            else:
                conn.execute("""
                    INSERT INTO Attendance(student_id,date,an_status,marked_by)
                    VALUES(?,?,?,?)
                    ON CONFLICT(student_id,date) DO UPDATE SET an_status=excluded.an_status
                """, (s["id"], date_str, "Ab", session["user_id"]))
                conn.execute("""
                    UPDATE Attendance SET an_status=? WHERE student_id=? AND date=?
                """, (status, s["id"], date_str))
            if status == "P":
                present_names.append(s["name"])

        conn.commit()
        update_excel(class_name, section, conn)
        summary = get_daily_summary(class_name, section, date_str, conn)
        conn.close()

        return render_template("ct/mark.html", success=True, detected=detected,
                               present=present_names, total=len(students),
                               date=date_str, session_type=session_type, summary=summary)

    return render_template("ct/mark.html")


# Date-wise Excel download
@app.route("/ct/excel/date")
@login_required(["class_teacher"])
def ct_excel_date():
    date_str = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))
    class_name = session["class_name"]
    section = session["section"]
    conn = get_conn()
    students = conn.execute(
        "SELECT id, name, roll_no FROM Students WHERE class_name=? AND section=? ORDER BY name ASC",
        (class_name, section)
    ).fetchall()
    att_rows = conn.execute(
        "SELECT student_id, fn_status, an_status FROM Attendance WHERE date=?",
        (date_str,)
    ).fetchall()
    conn.close()
    att_map = {r["student_id"]: r for r in att_rows}

    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = f"{class_name}-{section}-{date_str}"
    header = ["Roll No", "Name", f"{date_str} FN", f"{date_str} AN"]
    ws.append(header)
    for col in range(1, 5):
        c = ws.cell(1, col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1565C0")
        c.alignment = Alignment(horizontal="center")

    fn_p = an_p = 0
    for row_idx, s in enumerate(students, start=2):
        att = att_map.get(s["id"])
        fn = att["fn_status"] if att else "Ab"
        an = att["an_status"] if att else "Ab"
        ws.cell(row_idx, 1, s["roll_no"]).alignment = Alignment(horizontal="center")
        ws.cell(row_idx, 2, s["name"])
        fc = ws.cell(row_idx, 3, fn)
        fc.fill = PatternFill("solid", fgColor="C8E6C9" if fn == "P" else "FFCDD2")
        fc.alignment = Alignment(horizontal="center")
        ac = ws.cell(row_idx, 4, an)
        ac.fill = PatternFill("solid", fgColor="C8E6C9" if an == "P" else "FFCDD2")
        ac.alignment = Alignment(horizontal="center")
        if fn == "P": fn_p += 1
        if an == "P": an_p += 1

    # Summary row
    sr = len(students) + 2
    ws.cell(sr, 2, "Total Present").font = Font(bold=True, color="15803D")
    ws.cell(sr, 3, fn_p).font = Font(bold=True)
    ws.cell(sr, 3).alignment = Alignment(horizontal="center")
    ws.cell(sr, 4, an_p).font = Font(bold=True)
    ws.cell(sr, 4).alignment = Alignment(horizontal="center")
    ws.cell(sr+1, 2, "Total Absent").font = Font(bold=True, color="B91C1C")
    ws.cell(sr+1, 3, len(students)-fn_p).font = Font(bold=True)
    ws.cell(sr+1, 3).alignment = Alignment(horizontal="center")
    ws.cell(sr+1, 4, len(students)-an_p).font = Font(bold=True)
    ws.cell(sr+1, 4).alignment = Alignment(horizontal="center")

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"attendance_{class_name}_{section}_{date_str}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# Save camera captured image for enrollment
@app.route("/ct/enroll/camera", methods=["POST"])
@login_required(["class_teacher"])
def ct_enroll_camera():
    import base64
    data = request.json
    image_data = data.get("image", "")
    student_id = data.get("student_id")
    if not image_data or not student_id:
        return jsonify({"error": "Missing data"}), 400
    header, encoded = image_data.split(",", 1)
    img_bytes = base64.b64decode(encoded)
    class_name = session["class_name"]
    section = session["section"]
    folder = os.path.join(BASE_DIR, "dataset", class_name, section, f"student_{student_id}")
    os.makedirs(folder, exist_ok=True)
    existing = len([f for f in os.listdir(folder) if f.endswith(".jpg")])
    path = os.path.join(folder, f"img_{existing+1}.jpg")
    with open(path, "wb") as f:
        f.write(img_bytes)
    return jsonify({"success": True, "saved": existing+1})


# Save camera captured image for attendance
@app.route("/ct/mark/camera", methods=["POST"])
@login_required(["class_teacher"])
def ct_mark_camera():
    import base64
    data = request.json
    image_data = data.get("image", "")
    if not image_data:
        return jsonify({"error": "No image"}), 400
    header, encoded = image_data.split(",", 1)
    img_bytes = base64.b64decode(encoded)
    img_path = os.path.join(UPLOAD_DIR, "camera_capture.jpg")
    with open(img_path, "wb") as f:
        f.write(img_bytes)
    return jsonify({"success": True, "path": img_path})


# View & Edit Attendance
@app.route("/ct/attendance")
@login_required(["class_teacher"])
def ct_attendance():
    date_filter = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))
    conn = get_conn()
    records = conn.execute("""
        SELECT a.id, a.date, a.fn_status, a.an_status, a.subject, s.name, s.roll_no
        FROM Attendance a JOIN Students s ON a.student_id=s.id
        WHERE s.class_name=? AND s.section=? AND a.date=?
        ORDER BY s.name
    """, (session["class_name"], session["section"], date_filter)).fetchall()
    all_dates = conn.execute("""
        SELECT DISTINCT a.date FROM Attendance a JOIN Students s ON a.student_id=s.id
        WHERE s.class_name=? AND s.section=? ORDER BY a.date DESC
    """, (session["class_name"], session["section"])).fetchall()
    summary = get_daily_summary(session["class_name"], session["section"], date_filter, conn)
    conn.close()
    return render_template("ct/attendance.html", records=records,
                           date_filter=date_filter, all_dates=all_dates, summary=summary)


@app.route("/ct/attendance/edit/<int:aid>", methods=["POST"])
@login_required(["class_teacher"])
def ct_edit_attendance(aid):
    fn = request.form.get("fn_status", "Ab")
    an = request.form.get("an_status", "Ab")
    date = request.form.get("date", "")
    conn = get_conn()
    conn.execute("UPDATE Attendance SET fn_status=?, an_status=?, date=? WHERE id=?", (fn, an, date, aid))
    conn.commit()
    update_excel(session["class_name"], session["section"], conn)
    conn.close()
    return redirect(url_for("ct_attendance"))


@app.route("/ct/attendance/delete/<int:aid>", methods=["POST"])
@login_required(["class_teacher"])
def ct_delete_attendance(aid):
    conn = get_conn()
    conn.execute("DELETE FROM Attendance WHERE id=?", (aid,))
    conn.commit()
    update_excel(session["class_name"], session["section"], conn)
    conn.close()
    return redirect(url_for("ct_attendance"))


# Excel
@app.route("/ct/excel")
@login_required(["class_teacher"])
def ct_excel():
    path = os.path.join(EXCEL_DIR, f"{session['class_name']}_{session['section']}.xlsx")
    if os.path.exists(path):
        return send_file(path, as_attachment=True)
    return "No Excel file yet.", 404


# Timetable
@app.route("/ct/timetable", methods=["GET", "POST"])
@login_required(["class_teacher"])
def ct_timetable():
    if request.method == "POST":
        day = request.form.get("day", "")
        subject = request.form.get("subject", "")
        start = request.form.get("start_time", "")
        end = request.form.get("end_time", "")
        conn = get_conn()
        conn.execute(
            "INSERT INTO Timetable(teacher_id,day,subject,start_time,end_time) VALUES(?,?,?,?,?)",
            (session["user_id"], day, subject, start, end)
        )
        conn.commit()
        conn.close()
    conn = get_conn()
    entries = conn.execute(
        "SELECT * FROM Timetable WHERE teacher_id=? ORDER BY day, start_time",
        (session["user_id"],)
    ).fetchall()
    conn.close()
    return render_template("ct/timetable.html", entries=entries)


@app.route("/ct/timetable/delete/<int:tid>", methods=["POST"])
@login_required(["class_teacher"])
def ct_delete_timetable(tid):
    conn = get_conn()
    conn.execute("DELETE FROM Timetable WHERE id=? AND teacher_id=?", (tid, session["user_id"]))
    conn.commit()
    conn.close()
    return redirect(url_for("ct_timetable"))


# ── Student Portal ───────────────────────────────────────────────────
def student_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("student_class"):
            return redirect(url_for("student_login"))
        return f(*args, **kwargs)
    return wrapper


def _calc_attendance(students, att_rows, session_cols):
    att_map = {}
    for r in att_rows:
        for stype in ("FN", "AN"):
            key = f"{r['date']}_{stype}"
            att_map[(r["student_id"], key)] = r["fn_status"] if stype == "FN" else r["an_status"]
    result = []
    for s in students:
        total_present = sum(
            1 for col in session_cols
            if att_map.get((s["id"], col["key"]), "Ab") == "P"
        )
        total_sessions = len(session_cols)
        pct = round(total_present / total_sessions * 100) if total_sessions else 0
        result.append({
            "id": s["id"], "name": s["name"], "roll_no": s["roll_no"],
            "total_present": total_present, "total_sessions": total_sessions, "pct": pct,
            "att": {col["key"]: att_map.get((s["id"], col["key"]), "Ab") for col in session_cols}
        })
    return result


def _build_session_cols(att_rows):
    seen, cols = set(), []
    for r in sorted(att_rows, key=lambda x: x["date"]):
        for stype in ("FN", "AN"):
            key = f"{r['date']}_{stype}"
            if key not in seen:
                seen.add(key)
                cols.append({"date": r["date"], "stype": stype, "key": key})
    return cols


@app.route("/student", methods=["GET", "POST"])
@app.route("/student/login", methods=["GET", "POST"])
def student_login():
    conn = get_conn()
    classes = conn.execute("SELECT DISTINCT class_name FROM Students ORDER BY class_name").fetchall()
    sections = conn.execute("SELECT DISTINCT section FROM Students ORDER BY section").fetchall()
    conn.close()
    if request.method == "POST":
        year = request.form.get("year", "").strip()
        class_name = request.form.get("class_name", "").strip()
        section = request.form.get("section", "").strip()
        if not year or not class_name or not section:
            return render_template("student/login.html", classes=classes, sections=sections,
                                   error="Please fill all fields.",
                                   req_year=year, req_class=class_name, req_section=section)
        conn = get_conn()
        exists = conn.execute(
            "SELECT COUNT(*) as c FROM Students WHERE class_name=? AND section=?",
            (class_name, section)
        ).fetchone()["c"]
        conn.close()
        if not exists:
            return render_template("student/login.html", classes=classes, sections=sections,
                                   error="No students found for this class/section.",
                                   req_year=year, req_class=class_name, req_section=section)
        session["student_class"] = class_name
        session["student_section"] = section
        session["student_year"] = year
        return redirect(url_for("student_dashboard"))
    return render_template("student/login.html", classes=classes, sections=sections)


@app.route("/student/dashboard")
@student_required
def student_dashboard():
    class_name = session["student_class"]
    section = session["student_section"]
    date_filter = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))
    conn = get_conn()
    students = conn.execute(
        "SELECT id, name, roll_no FROM Students WHERE class_name=? AND section=? ORDER BY name",
        (class_name, section)
    ).fetchall()
    all_att = conn.execute("""
        SELECT a.student_id, a.date, a.fn_status, a.an_status
        FROM Attendance a JOIN Students s ON a.student_id=s.id
        WHERE s.class_name=? AND s.section=?
    """, (class_name, section)).fetchall()
    all_dates = conn.execute("""
        SELECT DISTINCT a.date FROM Attendance a JOIN Students s ON a.student_id=s.id
        WHERE s.class_name=? AND s.section=? ORDER BY a.date DESC
    """, (class_name, section)).fetchall()
    day_att = conn.execute(
        "SELECT a.student_id, a.fn_status, a.an_status FROM Attendance a "
        "JOIN Students s ON a.student_id=s.id "
        "WHERE s.class_name=? AND s.section=? AND a.date=?",
        (class_name, section, date_filter)
    ).fetchall()
    today_summary = get_daily_summary(class_name, section, date_filter, conn)
    total_days = len(set(r["date"] for r in all_att))
    conn.close()

    all_session_cols = _build_session_cols(all_att)
    day_map = {r["student_id"]: r for r in day_att}
    student_data = _calc_attendance(students, all_att, all_session_cols)
    records = []
    for s in student_data:
        day = day_map.get(s["id"])
        records.append({**s,
            "fn_status": day["fn_status"] if day else "Ab",
            "an_status": day["an_status"] if day else "Ab"})

    return render_template("student/dashboard.html",
        class_name=class_name, section=section, year=session["student_year"],
        records=records, date_filter=date_filter, all_dates=all_dates,
        total_students=len(students), total_days=total_days,
        today_fn_present=today_summary["fn_present"],
        today_fn_absent=today_summary["fn_absent"])


@app.route("/student/attendance")
@student_required
def student_attendance():
    class_name = session["student_class"]
    section = session["student_section"]
    conn = get_conn()
    students = conn.execute(
        "SELECT id, name, roll_no FROM Students WHERE class_name=? AND section=? ORDER BY name",
        (class_name, section)
    ).fetchall()
    all_att = conn.execute("""
        SELECT a.student_id, a.date, a.fn_status, a.an_status
        FROM Attendance a JOIN Students s ON a.student_id=s.id
        WHERE s.class_name=? AND s.section=?
    """, (class_name, section)).fetchall()
    conn.close()
    session_cols = _build_session_cols(all_att)
    student_data = _calc_attendance(students, all_att, session_cols)
    return render_template("student/attendance.html",
        class_name=class_name, section=section, year=session["student_year"],
        students=student_data, session_cols=session_cols)


@app.route("/student/excel")
@student_required
def student_excel():
    import io
    class_name = session["student_class"]
    section = session["student_section"]
    date_filter = request.args.get("date", "")
    conn = get_conn()
    students = conn.execute(
        "SELECT id, name, roll_no FROM Students WHERE class_name=? AND section=? ORDER BY name",
        (class_name, section)
    ).fetchall()
    all_att = conn.execute("""
        SELECT a.student_id, a.date, a.fn_status, a.an_status
        FROM Attendance a JOIN Students s ON a.student_id=s.id
        WHERE s.class_name=? AND s.section=?
    """, (class_name, section)).fetchall()
    conn.close()

    all_session_cols = _build_session_cols(all_att)
    student_data = _calc_attendance(students, all_att, all_session_cols)

    wb = Workbook()
    ws = wb.active

    if date_filter:
        ws.title = f"{class_name}-{section}-{date_filter}"
        header = ["Roll No", "Name", f"{date_filter} FN", f"{date_filter} AN",
                  "Total Present", "Total Sessions", "Attendance %"]
        ws.append(header)
        for col in range(1, len(header) + 1):
            c = ws.cell(1, col)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="0D9488")
            c.alignment = Alignment(horizontal="center")
        att_map = {r["student_id"]: r for r in all_att if r["date"] == date_filter}
        for ri, s in enumerate(student_data, start=2):
            att = att_map.get(s["id"])
            fn = att["fn_status"] if att else "Ab"
            an = att["an_status"] if att else "Ab"
            ws.cell(ri, 1, s["roll_no"]).alignment = Alignment(horizontal="center")
            ws.cell(ri, 2, s["name"])
            fc = ws.cell(ri, 3, fn)
            fc.fill = PatternFill("solid", fgColor="C8E6C9" if fn == "P" else "FFCDD2")
            fc.alignment = Alignment(horizontal="center")
            ac = ws.cell(ri, 4, an)
            ac.fill = PatternFill("solid", fgColor="C8E6C9" if an == "P" else "FFCDD2")
            ac.alignment = Alignment(horizontal="center")
            ws.cell(ri, 5, s["total_present"]).alignment = Alignment(horizontal="center")
            ws.cell(ri, 6, s["total_sessions"]).alignment = Alignment(horizontal="center")
            pc = ws.cell(ri, 7, f"{s['pct']}%")
            pc.fill = PatternFill("solid", fgColor="DCFCE7" if s["pct"] >= 75 else ("FEF3C7" if s["pct"] >= 60 else "FEE2E2"))
            pc.font = Font(bold=True)
            pc.alignment = Alignment(horizontal="center")
        fname = f"attendance_{class_name}_{section}_{date_filter}.xlsx"
    else:
        ws.title = f"{class_name}-{section}"
        header = ["Roll No", "Name"] + \
                 [f"{c['date']} {c['stype']}" for c in all_session_cols] + \
                 ["Total Present", "Total Sessions", "Attendance %"]
        ws.append(header)
        for col in range(1, len(header) + 1):
            c = ws.cell(1, col)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="0D9488")
            c.alignment = Alignment(horizontal="center")
        for ri, s in enumerate(student_data, start=2):
            ws.cell(ri, 1, s["roll_no"]).alignment = Alignment(horizontal="center")
            ws.cell(ri, 2, s["name"])
            for ci, col in enumerate(all_session_cols, start=3):
                status = s["att"].get(col["key"], "Ab")
                cell = ws.cell(ri, ci, status)
                cell.fill = PatternFill("solid", fgColor="C8E6C9" if status == "P" else "FFCDD2")
                cell.alignment = Alignment(horizontal="center")
            base = len(all_session_cols) + 3
            ws.cell(ri, base, s["total_present"]).alignment = Alignment(horizontal="center")
            ws.cell(ri, base + 1, s["total_sessions"]).alignment = Alignment(horizontal="center")
            pc = ws.cell(ri, base + 2, f"{s['pct']}%")
            pc.fill = PatternFill("solid", fgColor="DCFCE7" if s["pct"] >= 75 else ("FEF3C7" if s["pct"] >= 60 else "FEE2E2"))
            pc.font = Font(bold=True)
            pc.alignment = Alignment(horizontal="center")
        fname = f"attendance_{class_name}_{section}_full.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/student/logout")
def student_logout():
    session.pop("student_class", None)
    session.pop("student_section", None)
    session.pop("student_year", None)
    return redirect(url_for("student_login"))


if __name__ == "__main__":
    init_db()
    app.run(debug=True, port=5000)
