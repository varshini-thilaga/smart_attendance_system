import face_recognition
import pickle
import os
import sqlite3
import cv2
import pandas as pd
import numpy as np
from datetime import date, time, datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import shutil
import time as t

MODEL_FILE = "trained_model.pkl"

df = pd.read_csv("time_table.csv")

now = datetime.now()
current_day = date.today().weekday()
currentdate = now.strftime("%d_%m_%y")
duration = 0
roman = ['I', 'II', 'III', 'IV', 'V']

a = np.array(df.values[0][1:])
k = 0
start = []
end = []
periods = df.values[current_day + 1][1:]
dict_ = {}
for i in periods:
    dict_[a[k]] = i
    start.append(a[k][0:5])
    end.append(a[k][6:])
    k += 1


def is_time_between(begin_time, end_time, check_time=None):
    check_time = check_time or datetime.now().time()
    if begin_time < end_time:
        return begin_time <= check_time <= end_time
    return check_time >= begin_time or check_time <= end_time


def path12():
    global duration
    k = 0
    path = ""
    for i in dict_.keys():
        hs, ms = int(start[k][0:2]), int(start[k][3:5])
        he, me = int(end[k][0:2]), int(end[k][3:5])
        if is_time_between(time(hs, ms), time(he, me)):
            path = f"./{currentdate}/{dict_[i]}/{dict_[i]}_{currentdate}.xlsx"
            FMT = '%H:%M'
            tdelta = datetime.strptime(end[k], FMT) - datetime.strptime(start[k], FMT)
            duration = int(tdelta.seconds / 60 / 5)
            break
        k += 1
    return path


def detect():
    import dlib
    from PyQt5.QtWidgets import QApplication, QFileDialog
    import sys

    detector = dlib.get_frontal_face_detector()
    app = QApplication.instance() or QApplication(sys.argv)
    file, _ = QFileDialog.getOpenFileName(None, "Select Class Photo", "", "Images (*.jpg *.jpeg *.png)")

    if not file:
        print("No image selected.")
        return

    if not os.path.exists('./pics'):
        os.makedirs('./pics')
    img_src = cv2.imread(file)
    cv2.imwrite('./pics/framee1.jpg', img_src)

    img = cv2.imread('./pics/framee1.jpg')
    dets = detector(img, 1)

    if not os.path.exists('./Cropped_faces'):
        os.makedirs('./Cropped_faces')
    for f in os.listdir('./Cropped_faces'):
        os.remove(os.path.join('./Cropped_faces', f))

    for i, d in enumerate(dets):
        cv2.imwrite(f'./Cropped_faces/face{i+1}.jpg', img[d.top():d.bottom(), d.left():d.right()])
    print(f"Detected {len(dets)} faces")


def identify(path, column):
    if not os.path.exists(MODEL_FILE):
        print("Model not trained yet. Run train.py first.")
        return

    with open(MODEL_FILE, "rb") as f:
        data = pickle.load(f)

    known_encodings = data["encodings"]
    known_ids = data["ids"]

    connect = sqlite3.connect("Face-DataBase")
    wb = load_workbook(path)
    sheet = wb['Cse16']

    def getColumn():
        for i in range(1, len(list(sheet.rows)[0]) + 1):
            col = get_column_letter(i)
            if sheet[f'{col}1'].value == column:
                return col

    attend = {}

    for filename in os.listdir('./Cropped_faces'):
        if not filename.endswith(".jpg"):
            continue
        img_path = os.path.join('./Cropped_faces', filename)
        image = face_recognition.load_image_file(img_path)
        encodings = face_recognition.face_encodings(image)

        if not encodings:
            print(f"No face in {filename}")
            continue

        matches = face_recognition.compare_faces(known_encodings, encodings[0], tolerance=0.5)
        distances = face_recognition.face_distance(known_encodings, encodings[0])

        if True in matches:
            best_idx = np.argmin(distances)
            student_id = known_ids[best_idx]
            attend[student_id] = attend.get(student_id, 0) + 1
            print(f"Recognized student ID: {student_id}")
        else:
            print(f"Unknown face in {filename}")

    col = getColumn()
    for row in range(2, len(list(sheet.columns)[0]) + 1):
        rn = sheet.cell(row=row, column=1).value
        if rn is not None:
            rn_id = str(rn)[-3:]
            if rn_id in attend:
                sheet[f'{col}{row}'] = 1

    wb.save(path)
    connect.close()


def final_count(path):
    df = pd.read_excel(path, index_col=False)
    df.fillna(0, inplace=True)
    df = df.drop(0, axis=0)
    df['Final'] = df['I'] + df['II'] + df['III'] + df['IV'] + df['V']
    df['Final'] = np.where(df["Final"] >= 3, 'Present', 'Absent')
    df = df.drop(columns=['I', 'II', 'III', 'IV', 'V'])
    df.replace(0, np.nan, inplace=True)
    os.remove(path)
    df.to_excel(path)


def final_spreadsheet():
    df_col = {}
    for sub in os.listdir(currentdate):
        name = os.listdir(f"./{currentdate}/{sub}/")
        df_col[sub] = pd.read_excel(f"./{currentdate}/{sub}/{name[0]}", index_col=False)

    sub = list(df_col.keys())[-1]
    df_temp = df_col[sub].copy()
    df_temp['Roll Number'] = df_col[sub]['Roll Number']
    df_temp['Name'] = df_col[sub]['Name']

    for i in df_col.keys():
        df_temp[str(i)] = df_col[i]['Final']

    for i in os.listdir(f"./{currentdate}/"):
        shutil.rmtree(f"./{currentdate}/{i}")

    df_temp = df_temp.drop(["Unnamed: 0", "Final"], axis=1)
    df_temp.to_excel(f"./{currentdate}/{currentdate}_final.xlsx")


os.system("python spreadsheet.py")

interval = []
while True:
    if current_day != date.today().weekday() or currentdate != now.strftime("%d_%m_%y"):
        current_day = date.today().weekday()
        now = datetime.now()
        currentdate = now.strftime("%d_%m_%y")
        os.system("python spreadsheet.py")

    while is_time_between(time(int(start[0][0:2]), int(start[0][3:5])), time(int(end[-1][0:2]), int(end[-1][3:5]))):
        hello12 = datetime.now().time().strftime("%H:%M")
        curr = datetime.now()
        path = path12()

        if hello12 in start:
            interval = []
            for x in range(5):
                curr = curr + timedelta(minutes=duration)
                interval.append(curr.strftime("%H:%M"))
            t.sleep(60)

        if hello12 in interval:
            path = path12()
            index = interval.index(hello12)
            column12 = roman[index]
            detect()
            identify(path, column12)
            if column12 == "V":
                final_count(path)
            t.sleep(60)

        if hello12 == end[-1]:
            final_spreadsheet()
